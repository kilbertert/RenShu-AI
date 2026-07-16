"""SymMap v2 中医药知识图谱导入 (P1 阶段)

SymMap v2 公开数据（Wu et al. 2019），灌入 Neo4j。

节点（7 类，源 XLSX 主表，非 key file）：
  - Herb         (SMHB, 499 行)   Herb_id + 中/英/拼音/拉丁名 + 性味归经
  - TCMSymptom   (SMTS, 1717 行)  TCM_symptom_id + 中文/拼音名
  - MMSymptom    (SMMS, 961 行)   MM_symptom_id + UMLS_id / HPO_id
  - Ingredient   (SMIT, 19595 行) Ingredient_id + PubChem_id + OB_score
  - Target       (SMTT, 4302 行)  Target_id + Gene_symbol + NCBI/Uniprot
  - Disease      (SMDE, 5235 行)  Disease_id + OMIM_id / Orphanet_id
  - Syndrome     (SMSY, 233 行)   Syndrome_id + 中/英/拼音

关系：本脚本**不灌**。SymMap 关系需要从 supplementary tables 或
外部数据库（如 TCMSP/HERB 关联表）补充，单独写 import_symmap_v2_rels.py
或在 P1-3 改 _query_similar_syndromes 时一起处理。

依据：``判断.md`` §3.4 P1 范围 + 2026-06-08-p0.5-import-execution.md §3.4
执行::

    cd D:/AI/project/RenShu-AI/backend
    .venv/Scripts/python.exe -m scripts.import_symmap_v2 --db tcm_graph
    .venv/Scripts/python.exe -m scripts.import_symmap_v2 --db tcm_graph --dry-run
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

SYMMAP_DIR = Path("D:/AI/project/RenShu-AI/TCM_Dataset/SymMap_v2")
BATCH_SIZE = 500


@dataclass
class NodeTypeSpec:
    """单个节点类型的源文件 + 列映射"""
    label: str             # Neo4j label, e.g. "Herb"
    file_path: Path        # 绝对路径
    pk_col: str            # 主键列（XLSX 第一行表头中的字段名）
    field_map: dict[str, str]  # {xlsx_col -> neo4j_prop}（仅灌的字段，避免 Suppress 等噪声）
    suppress_col: str = "Suppress"  # 过滤掉被弃用的行（值非空则跳过）


@dataclass
class Stats:
    started_at: float = 0.0
    elapsed_sec: float = 0.0
    node_counts: dict[str, int] = field(default_factory=dict)
    skipped_suppressed: dict[str, int] = field(default_factory=dict)


# ===== 各类型规格定义 =====

HERB_SPEC = NodeTypeSpec(
    label="Herb",
    file_path=SYMMAP_DIR / "Herb" / "SymMap v2.0, SMHB file.xlsx",
    pk_col="Herb_id",
    field_map={
        "Herb_id": "id",
        "Chinese_name": "chinese_name",
        "Pinyin_name": "pinyin_name",
        "Latin_name": "latin_name",
        "English_name": "english_name",
        "Properties_Chinese": "properties_zh",
        "Properties_English": "properties_en",
        "Meridians_Chinese": "meridians_zh",
        "Meridians_English": "meridians_en",
        "Function": "function",
        "Class_Chinese": "class_zh",
        "Class_English": "class_en",
        "UsePart": "use_part",
    },
)

TCM_SYMPTOM_SPEC = NodeTypeSpec(
    label="TCMSymptom",
    file_path=SYMMAP_DIR / "TCM symptom" / "SymMap v2.0, SMTS file.xlsx",
    pk_col="TCM_symptom_id",
    field_map={
        "TCM_symptom_id": "id",
        "TCM_symptom_name": "name_zh",
        "Symptom_pinYin": "name_pinyin",  # 注意：文件里 PinYin 是大写 Y
        "Symptom_definition": "definition",
        "Symptom_locus": "locus",
        "Symptom_property": "property",
        "Type": "term_type",
    },
)

MM_SYMPTOM_SPEC = NodeTypeSpec(
    label="MMSymptom",
    file_path=SYMMAP_DIR / "MM symptom" / "SymMap v2.0, SMMS file.xlsx",
    pk_col="MM_symptom_id",
    field_map={
        "MM_symptom_id": "symmap_id",
        "MM_symptom_name": "name",
        "MM_symptom_definition": "definition",
        "UMLS_id": "umls_id",
        "HPO_id": "hpo_id",
        "OMIM_id": "omim_id",
        "ICD10CM_id": "icd10cm_id",
        "MeSH_id": "mesh_id",
        "MeSH_tree_numbers": "mesh_tree",
    },
)

INGREDIENT_SPEC = NodeTypeSpec(
    label="Ingredient",
    file_path=SYMMAP_DIR / "Ingredient" / "SymMap v2.0, SMIT file.xlsx",
    pk_col="Mol_id",  # 注意：文件里 PK 是 Mol_id，不是 Ingredient_id
    field_map={
        "Mol_id": "id",
        "Molecule_name": "molecule_name",
        "Molecule_formula": "molecule_formula",
        "Molecule_weight": "molecule_weight",
        "OB_score": "ob_score",
        "PubChem_CID": "pubchem_cid",  # 注意：PubChem_CID，不是 PubChem_id
        "CAS_id": "cas_id",
        "Type": "ingredient_type",
    },
)

TARGET_SPEC = NodeTypeSpec(
    label="Target",
    file_path=SYMMAP_DIR / "Target" / "SymMap v2.0, SMTT file.xlsx",
    pk_col="Gene_id",  # 注意：文件里 PK 是 Gene_id，不是 Target_id
    field_map={
        "Gene_id": "symmap_id",
        "Gene_symbol": "gene_symbol",
        "Gene_name": "gene_name",
        "Protein_name": "protein_name",
        "Chromosome": "chromosome",
        "NCBI_id": "ncbi_id",
        "UniProt_id": "uniprot_id",  # 注意：UniProt_id 大写 P，不是 Uniprot_id
        "Ensembl_id": "ensembl_id",
        "HGNC_id": "hgnc_id",
        "MIM_id": "omim_id",  # 注意：MIM_id 是 OMIM 风格
    },
)

DISEASE_SPEC = NodeTypeSpec(
    label="Disease",
    file_path=SYMMAP_DIR / "Disease" / "SymMap v2.0, SMDE file.xlsx",
    pk_col="Disease_id",
    field_map={
        "Disease_id": "symmap_id",
        "Disease_Name": "name",  # 注意：Disease_Name 大写 N
        "Disease_definition": "definition",
        "MeSH_id": "mesh_id",
        "OMIM_id": "omim_id",
        "Orphanet_id": "orphanet_id",
        "ICD10CM_id": "icd10cm_id",
        "UMLS_id": "umls_id",
        "MedDRA_id": "meddra_id",
    },
)

SYNDROME_SPEC = NodeTypeSpec(
    label="Syndrome",
    file_path=SYMMAP_DIR / "Syndrome" / "SymMap v2.0, SMSY file.xlsx",
    pk_col="Syndrome_id",
    field_map={
        "Syndrome_id": "id",
        "Syndrome_name": "name_zh",
        "Syndrome_PinYin": "name_pinyin",  # 注意：PinYin 大写 Y
        "Syndrome_English": "name_en",
        "Syndrome_definition": "definition",
        "Type": "term_type",
    },
)

ALL_SPECS: list[NodeTypeSpec] = [
    HERB_SPEC, TCM_SYMPTOM_SPEC, MM_SYMPTOM_SPEC, INGREDIENT_SPEC,
    TARGET_SPEC, DISEASE_SPEC, SYNDROME_SPEC,
]


# ===== 解析与统计 =====

def _is_suppressed(row: dict, spec: NodeTypeSpec) -> bool:
    """SymMap 的 Suppress 列：0 / None / 空串 → 保留；1 / 非零 → 跳过。
    实际值类型为 int（0/1），但历史上也见过字符串版，做防御性归一。"""
    val = row.get(spec.suppress_col)
    if val is None:
        return False
    if isinstance(val, str) and val.strip() in ("", "0"):
        return False
    if isinstance(val, (int, float)) and val == 0:
        return False
    return True


def _project_row(row: dict, spec: NodeTypeSpec) -> dict:
    """按 field_map 提取字段并重命名"""
    out = {}
    for src, dst in spec.field_map.items():
        v = row.get(src)
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        out[dst] = v
    return out


def iter_spec_rows(spec: NodeTypeSpec) -> Iterable[dict]:
    """生成 projected_row 流；suppressed 行以 {"__suppressed__": True} 表示"""
    import openpyxl
    wb = openpyxl.load_workbook(spec.file_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue
        d = dict(zip(headers, row))
        if _is_suppressed(d, spec):
            yield {"__suppressed__": True}
            continue
        projected = _project_row(d, spec)
        if spec.pk_col not in d or d[spec.pk_col] is None:
            continue
        projected[spec.field_map[spec.pk_col]] = d[spec.pk_col]
        yield projected


def collect_stats() -> Stats:
    """只统计，不写库"""
    s = Stats()
    print("[STATS] SymMap v2 节点规模探查 ...")
    for spec in ALL_SPECS:
        if not spec.file_path.exists():
            print(f"  [WARN] {spec.label}: 源文件不存在 {spec.file_path}")
            s.node_counts[spec.label] = 0
            continue
        n_kept = 0
        n_skip = 0
        for row in iter_spec_rows(spec):
            if "__suppressed__" in row:
                n_skip += 1
            else:
                n_kept += 1
        s.node_counts[spec.label] = n_kept
        s.skipped_suppressed[spec.label] = n_skip
        print(f"  {spec.label:12s}: {n_kept:6d} (跳过 Suppress={n_skip})")
    return s


# ===== 写入 =====

def _cypher_upsert(spec: NodeTypeSpec) -> str:
    """生成 MERGE Cypher；字段名经白名单 field_map 控制"""
    props = list(spec.field_map.values())
    pk_prop = props[0]
    extra = [f"n.{p} = row.{p}" for p in props[1:]]
    set_part = ", ".join(extra) if extra else ""
    cypher = f"""
    UNWIND $batch AS row
    MERGE (n:{spec.label} {{{pk_prop}: row.{pk_prop}}})
    {"SET " + set_part if set_part else ""}
    """
    return cypher.strip()


def write_to_neo4j(database: str) -> bool:
    try:
        from app.src.core.graph_db import get_neo4j_graph
    except ImportError as exc:
        print(f"[FAIL] 无法导入 graph_db: {exc}")
        return False

    os.environ["NEO4J_DB"] = database
    g = get_neo4j_graph()
    if g is None:
        print(f"[FAIL] Neo4j 未连接（database={database}）")
        return False

    total_start = time.monotonic()
    for spec in ALL_SPECS:
        if not spec.file_path.exists():
            continue
        cypher = _cypher_upsert(spec)
        batch: list[dict] = []
        t0 = time.monotonic()
        written = 0
        for row in iter_spec_rows(spec):
            if "__suppressed__" in row:
                continue
            batch.append(row)
            if len(batch) >= BATCH_SIZE:
                g.query(cypher, params={"batch": batch})
                written += len(batch)
                batch = []
        if batch:
            g.query(cypher, params={"batch": batch})
            written += len(batch)
        dt = time.monotonic() - t0
        print(f"  [WRITE] {spec.label:12s}: {written:6d} nodes ({dt:.1f}s)")
    print(f"[DONE] 全部写入完成，总耗时 {time.monotonic() - total_start:.1f}s")
    return True


def main(dry_run: bool = False, database: str = "tcm_graph") -> int:
    missing = [s.label for s in ALL_SPECS if not s.file_path.exists()]
    if missing:
        print(f"[FAIL] 以下节点源文件缺失: {missing}")
        return 1

    s = collect_stats()
    if dry_run:
        print("\n[DRY-RUN] 不写入 Neo4j")
        return 0

    print(f"\n[WRITE] -> Neo4j (database={database}) ...")
    return 0 if write_to_neo4j(database) else 1


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="只统计不写入")
    parser.add_argument("--db", default="tcm_graph",
                        choices=["tcm_graph", "neo4j"],
                        help="目标 Neo4j database (默认 tcm_graph)")
    args = parser.parse_args()
    raise SystemExit(main(dry_run=args.dry_run, database=args.db))
