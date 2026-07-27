"""ITCM 关系导入 (P2-6 阶段)

ITCM Manual Curation 两个 xlsx 的关系灌入 Neo4j。

关系（3 类，源文件）：

1. **Herb-Ingredient 关系** (HERB_CONTAINS_INGREDIENT)
   源：``Manual Curation of Herb Ingredient and Target.xlsx`` 总表
   规模：~106,618 行（每行 1 条关系，可能重复）
   节点匹配：HERB+(CHN) → Herb.name_zh
              INGREDIENT(ENG) → Ingredient.name

2. **Ingredient-Target 关系** (INGREDIENT_TARGETS)
   源：同上文件总表（related target 列）
   规模：~106,618 行（部分行有 gene symbol）
   节点匹配：INGREDIENT(ENG) → Ingredient.name
              related target (gene symbol) → Target.gene_symbol

3. **Formula-Herb 关系** (FORMULA_CONTAINS_HERB)
   源：``Manual Curation of Formula.xlsx`` ``formula-herb`` sheet
   规模：1,260 行
   节点匹配：方剂名 (CHN) → Formula.name_zh
              组成 (中药名) → Herb_ITCM.name_zh

执行::

    cd D:/AI/project/RenShu-AI/backend
    .venv/Scripts/python.exe -m scripts.import_itcm_rels --db neo4j
    .venv/Scripts/python.exe -m scripts.import_itcm_rels --db neo4j --dry-run
    .venv/Scripts/python.exe -m scripts.import_itcm_rels --db neo4j --only formula-herb
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, List

import openpyxl

BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from scripts.tcm_dataset_config import get_tcm_dataset_root

ITCM_DIR = get_tcm_dataset_root() / "ITCM"
BATCH_SIZE = 1000


@dataclass
class RelStats:
    rel_counts: dict[str, int] = field(default_factory=dict)
    skipped_unmatched: dict[str, int] = field(default_factory=dict)


# ===== 关系 1+2: Herb-Ingredient-Target 三元表 =====

HI_PATH = ITCM_DIR / "Manual Curation of Herb Ingredient and Target.xlsx"
HI_SHEET = "总表"


def iter_hi_rows():
    """从 Herb-Ingredient-Target 三元表 yield 行 dict"""
    wb = openpyxl.load_workbook(HI_PATH, read_only=True, data_only=True)
    ws = wb[HI_SHEET]
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if not any(c is not None for c in row):
            continue
        yield dict(zip(headers, row))


def extract_hi_relation_fields(row: dict) -> tuple[str, str, str]:
    """提取人工关系表的匹配键。

    ``ingredient_detail.txt`` 的 ``name`` 保存英文成分名，因此必须使用
    ``INGREDIENT(ENG)``。中文列只覆盖部分记录，并且无法匹配 ITCM 节点。
    """
    herb_zh = str(row.get("HERB+(CHN)") or "").strip()
    ingredient_name = str(row.get("INGREDIENT(ENG)") or "").strip()
    gene_symbol = str(row.get("related target (gene symbol)") or "").strip()
    return herb_zh, ingredient_name, gene_symbol


def collect_hi_stats() -> RelStats:
    """统计 HI 表中的关系（不去重，纯源行计数）"""
    s = RelStats()
    n_hi = 0
    n_it = 0
    for row in iter_hi_rows():
        herb_zh, ingredient_name, target = extract_hi_relation_fields(row)
        if herb_zh and ingredient_name:
            n_hi += 1
        if ingredient_name and target:
            n_it += 1
    s.rel_counts["HERB_CONTAINS_INGREDIENT"] = n_hi
    s.rel_counts["INGREDIENT_TARGETS"] = n_it
    return s


def write_hi_to_neo4j(database: str, dry_run: bool = False) -> dict:
    """写 Herb-Ingredient + Ingredient-Target 关系到 Neo4j"""
    try:
        from app.src.core.graph_db import get_neo4j_graph
    except ImportError as exc:
        print(f"[FAIL] 无法导入 graph_db: {exc}")
        return {}

    os.environ["NEO4J_DB"] = database
    g = get_neo4j_graph()
    if g is None:
        print(f"[FAIL] Neo4j 未连接（database={database}）")
        return {}

    cypher_hi = """
    UNWIND $batch AS row
    MATCH (h:Herb {source_db: 'ITCM', name_zh: row.herb_zh})
    MATCH (i:Ingredient {source_db: 'ITCM', name: row.ingredient_name})
    MERGE (h)-[r:HERB_CONTAINS_INGREDIENT]->(i)
    SET r.source = row.source,
        r.ingredient_category = row.category,
        r.pubchem_cid = coalesce(row.pubchem_cid, r.pubchem_cid)
    RETURN count(r) AS n
    """
    cypher_it = """
    UNWIND $batch AS row
    MATCH (i:Ingredient {source_db: 'ITCM', name: row.ingredient_name})
    MATCH (t:Target {source_db: 'ITCM', gene_symbol: row.gene_symbol})
    MERGE (i)-[r:INGREDIENT_TARGETS]->(t)
    SET r.source = row.source,
        r.pmid = row.pmid
    RETURN count(r) AS n
    """

    batch_hi: List[dict] = []
    batch_it: List[dict] = []
    written_hi = 0
    written_it = 0
    t0 = time.monotonic()
    n_rows = 0
    for row in iter_hi_rows():
        n_rows += 1
        herb_zh, ingredient_name, target = extract_hi_relation_fields(row)
        source_herb = (row.get("herb-ingredient source") or "").strip() if row.get("herb-ingredient source") else ""
        source_it = (row.get("ingredient-target source") or "").strip() if row.get("ingredient-target source") else ""
        category = (row.get("分类") or "").strip() if row.get("分类") else ""
        pmid_raw = row.get("pmids")
        pmid = None
        if pmid_raw is not None:
            pmid_str = str(pmid_raw).strip()
            if pmid_str:
                # 同一字段可能含多个 pmid（逗号分隔），取第一个
                first = pmid_str.split(",")[0].strip()
                try:
                    pmid = str(int(first))
                except ValueError:
                    pmid = first  # 非数字保留原值
        pubchem = row.get("pubchem cid")
        pubchem_cid = None
        if pubchem is not None:
            pubchem_str = str(pubchem).strip()
            if pubchem_str:
                first = pubchem_str.split(",")[0].strip()
                try:
                    pubchem_cid = str(int(float(first)))
                except ValueError:
                    pubchem_cid = first

        if herb_zh and ingredient_name:
            batch_hi.append({
                "herb_zh": herb_zh, "ingredient_name": ingredient_name,
                "pubchem_cid": pubchem_cid, "source": source_herb,
                "category": category,
            })
            if len(batch_hi) >= BATCH_SIZE:
                if not dry_run:
                    g.query(cypher_hi, params={"batch": batch_hi})
                written_hi += len(batch_hi)
                batch_hi = []

        if ingredient_name and target:
            batch_it.append({
                "ingredient_name": ingredient_name, "gene_symbol": target,
                "source": source_it, "pmid": pmid,
            })
            if len(batch_it) >= BATCH_SIZE:
                if not dry_run:
                    g.query(cypher_it, params={"batch": batch_it})
                written_it += len(batch_it)
                batch_it = []

    if batch_hi:
        if not dry_run:
            g.query(cypher_hi, params={"batch": batch_hi})
        written_hi += len(batch_hi)
    if batch_it:
        if not dry_run:
            g.query(cypher_it, params={"batch": batch_it})
        written_it += len(batch_it)

    dt = time.monotonic() - t0
    print(f"  [HI]   scanned {n_rows:7d} rows in {dt:.1f}s")
    print(f"  [HI]   HERB_CONTAINS_INGREDIENT : {written_hi:7d} source rows (MERGE 幂等)")
    print(f"  [HI]   INGREDIENT_TARGETS       : {written_it:7d} source rows (MERGE 幂等)")
    return {"herb_ingredient": written_hi, "ingredient_target": written_it}


# ===== 关系 3: Formula-Herb (formula-herb sheet) =====

FH_PATH = ITCM_DIR / "Manual Curation of Formula.xlsx"
FH_SHEET = "formula-herb"


def iter_fh_rows():
    """从 formula-herb sheet yield 行 dict"""
    wb = openpyxl.load_workbook(FH_PATH, read_only=True, data_only=True)
    ws = wb[FH_SHEET]
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if not any(c is not None for c in row):
            continue
        yield dict(zip(headers, row))


def collect_fh_stats() -> RelStats:
    s = RelStats()
    n = sum(1 for _ in iter_fh_rows())
    s.rel_counts["Formula-Herb_CONTAINS"] = n
    return s


def write_fh_to_neo4j(database: str, dry_run: bool = False) -> dict:
    """写 Formula-Herb 关系到 Neo4j"""
    try:
        from app.src.core.graph_db import get_neo4j_graph
    except ImportError as exc:
        print(f"[FAIL] 无法导入 graph_db: {exc}")
        return {}

    os.environ["NEO4J_DB"] = database
    g = get_neo4j_graph()
    if g is None:
        print(f"[FAIL] Neo4j 未连接（database={database}）")
        return {}

    cypher = """
    UNWIND $batch AS row
    MATCH (f:Formula {source_db: 'ITCM', name_zh: row.formula_name})
    MATCH (h:Herb {source_db: 'ITCM', name_zh: row.herb_name})
    MERGE (f)-[r:FORMULA_CONTAINS_HERB]->(h)
    RETURN count(r) AS n
    """

    batch: List[dict] = []
    written = 0
    t0 = time.monotonic()
    n_rows = 0
    for row in iter_fh_rows():
        n_rows += 1
        formula_name = (row.get("方剂名") or "").strip() if row.get("方剂名") else ""
        herb_name = (row.get("组成") or "").strip() if row.get("组成") else ""
        if not formula_name or not herb_name:
            continue
        batch.append({"formula_name": formula_name, "herb_name": herb_name})
        if len(batch) >= BATCH_SIZE:
            if not dry_run:
                g.query(cypher, params={"batch": batch})
            written += len(batch)
            batch = []
    if batch:
        if not dry_run:
            g.query(cypher, params={"batch": batch})
        written += len(batch)

    dt = time.monotonic() - t0
    print(f"  [FH]   scanned {n_rows:7d} rows in {dt:.1f}s")
    print(f"  [FH]   Formula-Herb_CONTAINS    : {written:7d} (MERGE 幂等)")
    return {"formula_herb": written}


# ===== 主流程 =====

def main(dry_run: bool = False, database: str = "neo4j", only: str | None = None) -> int:
    if not HI_PATH.exists():
        print(f"[FAIL] 缺失: {HI_PATH}")
        return 1
    if not FH_PATH.exists():
        print(f"[FAIL] 缺失: {FH_PATH}")
        return 1

    print(f"[STATS] ITCM 关系规模探查 ...")
    if only != "formula-herb":
        s_hi = collect_hi_stats()
        for k, v in s_hi.rel_counts.items():
            print(f"  {k:35s}: {v:7d}")
    if only != "hi":
        s_fh = collect_fh_stats()
        for k, v in s_fh.rel_counts.items():
            print(f"  {k:35s}: {v:7d}")

    if dry_run:
        print("\n[DRY-RUN] 不写入 Neo4j")
        return 0

    print(f"\n[WRITE] -> Neo4j (database={database}) ...")
    total_start = time.monotonic()
    if only != "formula-herb":
        print("\n[STAGE 1/2] Herb-Ingredient + Ingredient-Target ...")
        write_hi_to_neo4j(database, dry_run=dry_run)
    if only != "hi":
        print("\n[STAGE 2/2] Formula-Herb ...")
        write_fh_to_neo4j(database, dry_run=dry_run)
    print(f"\n[DONE] 全部写入完成，总耗时 {time.monotonic() - total_start:.1f}s")
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="只统计不写入")
    parser.add_argument("--db", default="neo4j",
                        choices=["tcm_graph", "neo4j"],
                        help="目标 Neo4j database (默认 neo4j)")
    parser.add_argument("--only", choices=["hi", "formula-herb"], default=None,
                        help="只跑其中一个阶段（调试用）")
    args = parser.parse_args()
    raise SystemExit(main(dry_run=args.dry_run, database=args.db, only=args.only))
