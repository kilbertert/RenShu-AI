"""TCMBank 中医药数据库节点导入 (P2 阶段)

TCMBank 公开数据（3 个 xlsx + 41,812 个 mol2 分子结构文件）。

本脚本**仅灌节点**（3 个 xlsx 主表），不处理 mol2 文件
（mol2 是分子结构描述符，需 RDKit 解析 + 化学指纹，不属于图谱节点范围；
若有需求可单独写 import_tcmbank_mol2.py）。

节点（3 类，源 xlsx）：
  - Disease_TCMBank    (disease_all.xlsx,    32,530 行)  Disease_id + DisGENet/HPO/MeSH 桥接
  - Target_TCMBank     (gene_all.xlsx,       15,111 行)  Target_id + Gene_name + TTD
  - Ingredient_TCMBank (ingredient_all.xlsx, 61,966 行)  TCMBank_ID + Smiles + SymMap 桥接
                                                         （253 列仅取关键 20 列，避免属性爆炸）

依据：``判断.md`` §3.4 P2 范围
执行::

    cd D:/AI/project/RenShu-AI/backend
    .venv/Scripts/python.exe -m scripts.import_tcmbank --db neo4j
    .venv/Scripts/python.exe -m scripts.import_tcmbank --db neo4j --dry-run
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

TCMBANK_DIR = Path("D:/AI/project/RenShu-AI/TCM_Dataset/TCMBank")
BATCH_SIZE = 500


@dataclass
class NodeTypeSpec:
    label: str
    file_path: Path
    pk_col: str
    field_map: dict[str, str]  # {xlsx_col -> neo4j_prop}
    skip_blank_pk: bool = True


@dataclass
class Stats:
    node_counts: dict[str, int] = field(default_factory=dict)
    skipped_blank: dict[str, int] = field(default_factory=dict)


# ===== 各类型规格定义 =====

DISEASE_SPEC = NodeTypeSpec(
    label="Disease_TCMBank",
    file_path=TCMBANK_DIR / "disease_all.xlsx",
    pk_col="Disease_id",
    field_map={
        "Disease_id": "tcmbank_id",
        "Disease_name": "name",
        "Disease_type": "disease_type",
        "DisGENet_disease_id": "disgenet_id",
        "DiseaseClass_MeSH": "mesh_class_id",
        "DiseaseClassName_MeSH": "mesh_class_name",
        "HPO_ClassId": "hpo_class_id",
        "HPO_ClassName": "hpo_class_name",
        "DO_ClassId": "do_class_id",
        "DO_ClassName": "do_class_name",
        "UMLS_SemanticTypeId": "umls_semantic_type_id",
        "UMLS_SemanticTypeName": "umls_semantic_type_name",
        "Source_ID": "source_id",
    },
)

TARGET_SPEC = NodeTypeSpec(
    label="Target_TCMBank",
    file_path=TCMBANK_DIR / "gene_all.xlsx",
    pk_col="Target_id",
    field_map={
        "Target_id": "tcmbank_id",
        "Gene_name": "gene_name",
        "Gene_alias": "gene_alias",
        "Chromosome": "chromosome",
        "Map_location": "map_location",
        "Description": "description",
        "Type_of_gene": "gene_type",
        "TTD_target_id": "ttd_target_id",
        "TTD_target_type": "ttd_target_type",
        "Source_ID": "source_id",
    },
)

# ingredient_all.xlsx 253 列，只取核心 25 列（基础标识 + 桥接 + 关键化学指标）
# 完整化学描述符留给 RDKit / mol2 解析脚本
INGREDIENT_SPEC = NodeTypeSpec(
    label="Ingredient_TCMBank",
    file_path=TCMBANK_DIR / "ingredient_all.xlsx",
    pk_col="TCMBank_ID",
    field_map={
        "TCMBank_ID": "tcmbank_id",
        "Ingredient_id": "ingredient_id",
        "name": "name",
        "Alias": "alias",
        "Molecular_Formula": "molecular_formula",
        "Molecular_Weight": "molecular_weight",
        "Molecular_Volume": "molecular_volume",
        "ALogP": "alogp",
        "Molecular_PolarSurfaceArea": "psa",
        "Smiles": "smiles",
        "OB_score": "ob_score",
        "CAS_id": "cas_id",
        "SymMap_id": "symmap_id",
        "TCMID_id": "tcmid_id",
        "TCMSP_id": "tcmsp_id",
        "TCM-ID_id": "tcm_id_id",
        "PubChem_id": "pubchem_id",
        "DrugBank_id": "drugbank_id",
        "level1_name": "taxonomy_l1_zh",
        "level1_name_en": "taxonomy_l1_en",
        "level2_name": "taxonomy_l2_zh",
        "level2_name_en": "taxonomy_l2_en",
        "TCM_name": "tcm_name_zh",
        "TCM_name_en": "tcm_name_en",
        "TCM_name2": "tcm_name2",
    },
)

ALL_SPECS: list[NodeTypeSpec] = [DISEASE_SPEC, TARGET_SPEC, INGREDIENT_SPEC]


# ===== 解析 =====

def _project_row(headers: list[str], values: list, spec: NodeTypeSpec) -> dict | None:
    d = dict(zip(headers, values))
    pk_val = d.get(spec.pk_col)
    if pk_val is None or (isinstance(pk_val, str) and pk_val.strip() == ""):
        return None
    out = {}
    for src, dst in spec.field_map.items():
        v = d.get(src)
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        out[dst] = v
    return out


def iter_spec_rows(spec: NodeTypeSpec) -> Iterable[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(spec.file_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue
        projected = _project_row(headers, list(row), spec)
        if projected is None:
            yield {"__blank_pk__": True}
        else:
            yield projected


def collect_stats() -> Stats:
    s = Stats()
    print("[STATS] TCMBank 节点规模探查 ...")
    for spec in ALL_SPECS:
        if not spec.file_path.exists():
            print(f"  [WARN] {spec.label}: 源文件不存在 {spec.file_path}")
            s.node_counts[spec.label] = 0
            continue
        n_kept = 0
        n_skip = 0
        for row in iter_spec_rows(spec):
            if "__blank_pk__" in row:
                n_skip += 1
            else:
                n_kept += 1
        s.node_counts[spec.label] = n_kept
        s.skipped_blank[spec.label] = n_skip
        print(f"  {spec.label:20s}: {n_kept:7d} (跳过空 PK={n_skip})")
    return s


# ===== 写入 =====

def _cypher_upsert(spec: NodeTypeSpec) -> str:
    props = list(spec.field_map.values())
    pk_prop = props[0]
    extra = [f"n.{p} = row.{p}" for p in props[1:]]
    set_part = ", ".join(extra) if extra else ""
    cypher = f"""
    UNWIND $batch AS row
    MERGE (n:{spec.label} {{{pk_prop}: row.{pk_prop}}})
    {"SET " + set_part if set_part else ""}
    """
    return cypher.strip()


def write_to_neo4j(database: str) -> bool:
    try:
        from app.src.core.graph_db import get_neo4j_graph
    except ImportError as exc:
        print(f"[FAIL] 无法导入 graph_db: {exc}")
        return False

    os.environ["NEO4J_DB"] = database
    g = get_neo4j_graph()
    if g is None:
        print(f"[FAIL] Neo4j 未连接（database={database}）")
        return False

    total_start = time.monotonic()
    for spec in ALL_SPECS:
        if not spec.file_path.exists():
            continue
        cypher = _cypher_upsert(spec)
        batch: list[dict] = []
        t0 = time.monotonic()
        written = 0
        for row in iter_spec_rows(spec):
            if "__blank_pk__" in row:
                continue
            batch.append(row)
            if len(batch) >= BATCH_SIZE:
                g.query(cypher, params={"batch": batch})
                written += len(batch)
                batch = []
        if batch:
            g.query(cypher, params={"batch": batch})
            written += len(batch)
        dt = time.monotonic() - t0
        print(f"  [WRITE] {spec.label:20s}: {written:7d} nodes ({dt:.1f}s)")
    print(f"[DONE] 全部写入完成，总耗时 {time.monotonic() - total_start:.1f}s")
    return True


def main(dry_run: bool = False, database: str = "neo4j") -> int:
    missing = [s.label for s in ALL_SPECS if not s.file_path.exists()]
    if missing:
        print(f"[FAIL] 以下节点源文件缺失: {missing}")
        return 1

    s = collect_stats()
    if dry_run:
        print("\n[DRY-RUN] 不写入 Neo4j")
        return 0

    print(f"\n[WRITE] -> Neo4j (database={database}) ...")
    return 0 if write_to_neo4j(database) else 1


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="只统计不写入")
    parser.add_argument("--db", default="neo4j",
                        choices=["tcm_graph", "neo4j"],
                        help="目标 Neo4j database (默认 neo4j)")
    args = parser.parse_args()
    raise SystemExit(main(dry_run=args.dry_run, database=args.db))
