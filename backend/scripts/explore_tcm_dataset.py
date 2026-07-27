"""TCM_Dataset 数据探查脚本（v2）

遍历 ``D:/AI/project/RenShu-AI/TCM_Dataset/`` 下三个数据库的目录结构，
输出每个数据文件的**编码、行数、字段、样例**到
``backend/docs/tcm_dataset_exploration_report.md``，并输出一份**字段映射
初稿**到 ``backend/docs/tcm_dataset_field_mapping.md``。

依据：``判断.md`` §2 P0 探查阶段。
v2 调整：实际数据是 ``.xlsx`` / ``.txt``，蓝图写的 CSV/TSV 假设需要改写。

执行::

    cd D:/AI/project/RenShu-AI/backend
    python -m scripts.explore_tcm_dataset

输出：
    backend/docs/tcm_dataset_exploration_report.md
    backend/docs/tcm_dataset_field_mapping.md
"""
from __future__ import annotations

import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from scripts.tcm_dataset_config import get_tcm_dataset_root

TCM_ROOT = get_tcm_dataset_root()
REPORT_PATH = BACKEND_ROOT / "docs" / "tcm_dataset_exploration_report.md"
MAPPING_PATH = BACKEND_ROOT / "docs" / "tcm_dataset_field_mapping.md"

SKIP_DIRS = {"all_mol2"}
SUPPORTED_SUFFIXES = (".csv", ".tsv", ".txt", ".xlsx", ".obo", ".hpoa")

# 文件 → 节点标签 的初稿（基于 README 推断，P0-3 人工审阅后纠正）
FILE_TO_NODE_HINT = {
    "SymMap v2.0, SMHB file.xlsx": "Herb",
    "SymMap v2.0, SMTS file.xlsx": "TCMSymptom",
    "SymMap v2.0, SMMS file.xlsx": "MMSymptom",
    "SymMap v2.0, SMIT file.xlsx": "Ingredient",
    "SymMap v2.0, SMTT file.xlsx": "Target",
    "SymMap v2.0, SMDE file.xlsx": "Disease",
    "SymMap v2.0, SMSY file.xlsx": "Syndrome",
    "ingredient_all.xlsx": "Ingredient (TCMBank)",
    "gene_all.xlsx": "Target (TCMBank)",
    "disease_all.xlsx": "Disease (TCMBank)",
    "formula_detail.txt": "Formula",
    "herb_detail.txt": "Herb (ITCM)",
    "ingredient_detail.txt": "Ingredient (ITCM)",
    "target_detail.txt": "Target (ITCM)",
    "disease_detail.txt": "Disease (ITCM)",
    "Manual Curation of Formula.xlsx": "Formula 人工关系",
    "Manual Curation of Herb Ingredient and Target.xlsx": "Herb->Ing / Ing->Target 人工关系",
    "Network used in TOOL.txt": "Network (PPI/Drug-Target)",
    "Network id used in TOOL.txt": "Network (ID索引)",
    "Drugbank and Its Targets_Used in TOOL.txt": "Drug->Target 参考",
    "6878 Global Approved Small Molecules and Their Category Properties Prediction.xlsx": "Reference Drug (不导入)",
}


@dataclass
class FileInfo:
    path: Path
    db_name: str
    sub_name: str
    suffix: str
    encoding: Optional[str] = None
    row_count: int = 0
    fields: list[str] = field(default_factory=list)
    sample: list[str] = field(default_factory=list)
    error: Optional[str] = None
    file_size_kb: float = 0.0


def detect_text_encoding(path: Path) -> str:
    """GBK/UTF-8 兜底（用于 .txt/.csv/.tsv）。"""
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            with open(path, encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "unknown"


def read_xlsx_info(path: Path) -> FileInfo:
    """读取 .xlsx 第一个 sheet 的行数、字段、第一行数据。"""
    from openpyxl import load_workbook

    info = FileInfo(
        path=path,
        db_name="",
        sub_name="",
        suffix=".xlsx",
        file_size_kb=path.stat().st_size / 1024,
    )
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            info.error = "空 sheet"
            return info
        info.fields = [str(h) if h is not None else "" for h in header]
        first_data = next(rows, None)
        if first_data:
            info.sample = [str(v) if v is not None else "" for v in first_data]
        info.row_count = sum(1 for _ in rows) + (1 if first_data else 0)
    except Exception as exc:
        info.error = f"{type(exc).__name__}: {exc}"
    return info


def read_text_info(path: Path) -> FileInfo:
    """读取 .txt/.csv/.tsv：先探编码、推断分隔符、读 header + 一行样例。"""
    info = FileInfo(
        path=path,
        db_name="",
        sub_name="",
        suffix=path.suffix,
        file_size_kb=path.stat().st_size / 1024,
    )
    enc = detect_text_encoding(path)
    info.encoding = enc
    if enc == "unknown":
        info.error = "无法识别编码"
        return info

    with open(path, encoding=enc, errors="replace") as f:
        first_line = next((ln.rstrip("\n\r") for ln in f if ln.strip()), None)
        if first_line is None:
            info.error = "空文件"
            return info
        second_line = next((ln.rstrip("\n\r") for ln in f if ln.strip()), None)
        remaining = sum(1 for ln in f if ln.strip())

    header = first_line
    sep = "\t" if "\t" in header else ("," if "," in header else ("|" if "|" in header else None))
    if sep is None:
        info.fields = [header]
        info.sample = second_line.split() if second_line else []
        info.row_count = (1 if second_line else 0) + remaining
        return info

    info.fields = [c.strip() for c in header.split(sep)]
    info.row_count = (1 if second_line else 0) + remaining
    if second_line:
        info.sample = [c.strip() for c in second_line.split(sep)]
    return info


def explore_file(path: Path, db_name: str, sub_name: str) -> FileInfo:
    if path.suffix.lower() == ".xlsx":
        info = read_xlsx_info(path)
    else:
        info = read_text_info(path)
    info.db_name = db_name
    info.sub_name = sub_name
    return info


def iter_tcm_files() -> list[FileInfo]:
    """按 (db, sub) 维度收集所有待探查文件。"""
    results: list[FileInfo] = []
    for db_dir in sorted(TCM_ROOT.iterdir()):
        if not db_dir.is_dir():
            continue
        for entry in sorted(db_dir.iterdir()):
            if entry.is_file() and entry.suffix.lower() in SUPPORTED_SUFFIXES:
                results.append(explore_file(entry, db_dir.name, "<root>"))
            elif entry.is_dir() and entry.name not in SKIP_DIRS:
                for f in sorted(entry.iterdir()):
                    if f.is_file() and f.suffix.lower() in SUPPORTED_SUFFIXES:
                        results.append(explore_file(f, db_dir.name, entry.name))
    return results


def render_report(infos: list[FileInfo]) -> str:
    """渲染 markdown 探查报告。"""
    lines: list[str] = ["# TCM_Dataset 探查报告 (v2)\n"]
    lines.append("> 自动生成于 `scripts/explore_tcm_dataset.py`\n")
    lines.append(f"> 数据源：`{TCM_ROOT}`\n")
    lines.append(
        f"> **统计**：{len(infos)} 个数据文件，"
        f"{sum(1 for i in infos if not i.error)} 个可读，"
        f"{sum(1 for i in infos if i.error)} 个失败\n\n"
    )

    by_db: dict[str, list[FileInfo]] = {}
    for info in infos:
        by_db.setdefault(info.db_name, []).append(info)

    for db_name in sorted(by_db):
        lines.append(f"## {db_name}\n\n")
        for info in by_db[db_name]:
            rel = info.path.relative_to(TCM_ROOT)
            if info.error:
                lines.append(f"### [FAIL] `{rel}`\n\n- 错误: {info.error}\n\n")
                continue
            lines.append(f"### `{rel}`\n\n")
            lines.append(f"- **大小**: {info.file_size_kb:.1f} KB\n")
            lines.append(f"- **格式**: `{info.suffix.lstrip('.')}`\n")
            if info.encoding:
                lines.append(f"- **编码**: `{info.encoding}`\n")
            lines.append(f"- **行数**: {info.row_count}\n")
            lines.append(f"- **字段数**: {len(info.fields)}\n")
            lines.append(
                "- **字段**: "
                + ", ".join(f"`{f}`" for f in info.fields[:12])
                + (f" ... (+{len(info.fields) - 12})" if len(info.fields) > 12 else "")
                + "\n"
            )
            if info.sample:
                sample_str = " | ".join(info.sample[:6])
                if len(info.sample) > 6:
                    sample_str += f" ... (+{len(info.sample) - 6})"
                lines.append(f"- **样例**: `{sample_str}`\n")
            lines.append("\n")

    lines.append("---\n\n## 探查脚本调用记录\n\n")
    lines.append("```\n")
    lines.append("python -m scripts.explore_tcm_dataset\n")
    lines.append("```\n")
    return "".join(lines)


def render_field_mapping(infos: list[FileInfo]) -> str:
    """渲染字段映射初稿（人工审阅版）。"""
    lines: list[str] = ["# TCM_Dataset 字段映射表 (v2 初稿)\n"]
    lines.append("> 由 `explore_tcm_dataset.py` 自动生成初稿，**需 P0-3 人工审阅签字**\n")
    lines.append("> 状态：`[TODO]` = 蓝图假设未核实；`[OK]` = 已人工核对\n\n")
    lines.append("## 全局约定\n\n")
    lines.append("- **跨库节点用前缀 ID** 避免冲突：\n")
    lines.append("  - SymMap_v2 -> `SM_`\n")
    lines.append("  - TCMBank -> `TCMB_`\n")
    lines.append("  - ITCM -> `ITCM_`\n")
    lines.append("- **MERGE 而非 CREATE** 保持幂等\n")
    lines.append("- **UNWIND + batch_size=1000** 批量写入\n")
    lines.append("- **xlsx 文件** 通过 openpyxl 读第一 sheet；`Key file` 通常作为全文检索索引，本次灌库不直接导入（导入 `file` 即可）\n\n")

    for info in sorted(infos, key=lambda i: (i.db_name, i.path)):
        rel = info.path.relative_to(TCM_ROOT)
        hint = FILE_TO_NODE_HINT.get(info.path.name, "[TODO] 待 P0-3 确认")
        status = "[TODO] 待审" if hint.startswith("[TODO]") else "[OK] 已映射"
        lines.append(f"## `{rel}`\n\n")
        lines.append(f"- **归属（自动推断）**: {hint}\n")
        lines.append(f"- **状态**: {status}\n")
        if info.error:
            lines.append(f"- **错误**: {info.error}\n\n")
            continue
        lines.append(f"- **行数**: {info.row_count}\n")
        lines.append(f"- **字段**: `{', '.join(info.fields)}`\n")
        lines.append(f"- **目标 Neo4j 节点标签（建议）**: \n")
        lines.append(f"  - `{hint.split()[0] if not hint.startswith('[TODO]') else '?'}` （具体待 P0-3 确认）\n")
        lines.append(f"- **属性映射**: 见 `import_<source>.py` 中 NodeSpec.extra_fields\n\n")

    return "".join(lines)


def main() -> int:
    if not TCM_ROOT.exists():
        print(f"[FAIL] TCM_Dataset 根目录不存在: {TCM_ROOT}")
        return 1

    print(f"[DIR] 探查 {TCM_ROOT} ...")
    infos = iter_tcm_files()
    print(f"   发现 {len(infos)} 个数据文件")

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(render_report(infos), encoding="utf-8")
    MAPPING_PATH.write_text(render_field_mapping(infos), encoding="utf-8")
    print(f"[OK] 报告: {REPORT_PATH}")
    print(f"[OK] 字段映射: {MAPPING_PATH}")
    print(f"   跳过目录: {SKIP_DIRS}")
    print(f"   失败: {sum(1 for i in infos if i.error)} 个")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
